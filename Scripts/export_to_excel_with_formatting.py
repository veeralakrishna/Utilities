import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.formatting.rule import DataBarRule


def export_to_excel_with_formatting(df, conditional_cols, delta_cols, conditional_method, wrap_cols, output_file):
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    
    # Convert DataFrame to Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Apply center align and middle align to all cells
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # Apply all borders for all cells
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            # Format header row
            if r_idx == 1:
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                cell.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
                # Apply thick border to header cells
                cell.border = Border(bottom=Side(style='medium'))
            # Apply conditional formatting
            elif df.columns[c_idx - 1] in conditional_cols:
                if conditional_method == 'color_scale':
                    if isinstance(value, (int, float)):
                        min_val = df[df.columns[c_idx - 1]].min()
                        max_val = df[df.columns[c_idx - 1]].max()
                        scale = (value - min_val) / (max_val - min_val)
                        r = int(255 * (1 - scale))
                        g = int(255 * scale)
                        b = 200
                        color_hex = f'{r:02x}{g:02x}{b:02x}'
                        cell.font = Font(color=color_hex, bold=True)
            # Apply formatting to delta columns
            elif df.columns[c_idx - 1] in delta_cols:
                red_arrow = u'\u2193'
                green_arrow = u'\u2191'
                yellow_arrow = u'\u2195'
                if cell.value > 0:
                    cell.value = f"{cell.value} {green_arrow}"
                    cell.font = Font(color="008000")
                elif cell.value < 0:
                    cell.value = f"{cell.value} {red_arrow}"
                    cell.font = Font(color="FF0000")
                else:
                    cell.value = f"{cell.value} {yellow_arrow}"
                    cell.font = Font(color="FFFF00")
            # Apply wrap text to specified columns
            if df.columns[c_idx - 1] in wrap_cols:
                cell.alignment = cell.alignment.copy(wrap_text=True)
                ws.column_dimensions[cell.column_letter].width = 50
                    
    # Autofit column width
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 50:
            adjusted_width = 50
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
    # Save the workbook
    wb.save(output_file)

# Example usage:
# conditional_cols = ['Column1', 'Column2']
# delta_cols = ['Column3', 'Column4']
# wrap_cols = ['Column2', 'Column4']
# df = pd.DataFrame({'Column1': [1, -2, 0], 'Column2': [-1, 2, 0], 'Column3': [10, -5, 0], 'Column4': [5, -10, 0]})
# export_to_excel_with_formatting(df, conditional_cols, delta_cols, 'color_scale', wrap_cols, 'formatted_data.xlsx')
